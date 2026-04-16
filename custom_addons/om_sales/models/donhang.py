from odoo import api, fields, models
from odoo.exceptions import UserError


class ShoppingCart(models.Model):
  _name = 'sm.shopping.cart'
  _description = 'Giỏ hàng'
  _inherit = ['mail.thread']

  name = fields.Char(string='Mã đơn hàng', copy=False, readonly=True, default='Mới')

  date_order = fields.Datetime(string='Ngày đặt', default=fields.Datetime.now)
  customer_name = fields.Char(string='Tên khách hàng', )
  customer_phone = fields.Char(string='Số điện thoại', )
  customer_email = fields.Char(string='Email')
  customer_address = fields.Text(string='Địa chỉ giao hàng')

  cart_line_ids = fields.One2many('sm.shopping.cart.line', 'cart_id', string='Danh sách sản phẩm')

  coupon_id = fields.Many2one('sm.coupon', string='Mã giảm giá (tương thích cũ)')
  coupon_ids = fields.Many2many('sm.coupon', 'sm_shopping_cart_coupon_rel', 'cart_id', 'coupon_id',
                                string='Mã giảm giá đã áp dụng')
  discount_amount = fields.Float(string='Số tiền giảm', default=0.0)

  subtotal_price = fields.Float(string='Tạm tính', compute='_compute_total_price', store=True)

  total_price = fields.Float(string='Tổng tiền', compute='_compute_total_price', store=True)

  payment_type = fields.Selection([('cod', 'COD'), ('bank_transfer', 'Chuyển khoản')], string='Hình thức thanh toán',
                                  default='cod')

  state = fields.Selection([('draft', 'Nháp'), ('awaiting_confirmation', 'Chờ xác nhận'), ('confirmed', 'Đã xác nhận'),
                            ('shipping', 'Đang giao hàng'), ('done', 'Hoàn thành'), ('returned', 'Khách trả hàng'),
                            ('cancel', 'Đã hủy')], string='Trạng thái', default='draft', tracking=True, copy=False,
                           group_expand='_read_group_state')

  picking_ids = fields.One2many('sm.stock.picking', 'cart_id', string='Các Phiếu Kho')
  picking_count = fields.Integer(string='Số lượng Phiếu', compute='_compute_picking_count')

  @api.model
  def _read_group_state(self, states, domain, order):
    return ['draft', 'awaiting_confirmation', 'confirmed', 'shipping', 'done', 'returned']

  @api.depends('picking_ids')
  def _compute_picking_count(self):
    for rec in self:
      rec.picking_count = len(rec.picking_ids)

  def action_view_pickings(self):
    self.ensure_one()
    action = self.env.ref('om_sales.action_sm_stock_picking').read()[0]
    pickings = self.picking_ids
    if len(pickings) > 1:
      action['domain'] = [('id', 'in', pickings.ids)]
    elif pickings:
      action['views'] = [(self.env.ref('om_sales.view_sm_stock_picking_form').id, 'form')]
      action['res_id'] = pickings.id
    return action

  def _get_default_locations(self):
    internal_loc = self.env['sm.stock.location'].search([('location_type', '=', 'internal')], limit=1)
    if not internal_loc:
      internal_loc = self.env['sm.stock.location'].create({'name': 'Kho Mặc Định', 'location_type': 'internal'})
    customer_loc = self.env['sm.stock.location'].search([('location_type', '=', 'customer')], limit=1)
    if not customer_loc:
      customer_loc = self.env['sm.stock.location'].create({'name': 'Kho Khách Hàng', 'location_type': 'customer'})
    return internal_loc, customer_loc

  @api.depends('cart_line_ids.quantity', 'cart_line_ids.price_unit', 'discount_amount')
  def _compute_total_price(self):
    for rec in self:
      subtotal = sum(line.quantity * line.price_unit for line in rec.cart_line_ids)
      rec.subtotal_price = subtotal
      total = subtotal - rec.discount_amount
      rec.total_price = total if total > 0 else 0
      print(
        f"DEBUG_COMPUTE: Cart {rec.id}, Subtotal: {subtotal}, Discount: {rec.discount_amount}, Final: {rec.total_price}")

  @api.model
  def create(self, vals):
    print(f"DEBUG_MODEL_CREATE: Incoming vals: {vals}")
    if vals.get('name', 'Mới') == 'Mới':
      vals['name'] = self.env['ir.sequence'].next_by_code('sm.shopping.cart') or 'GH0001'
    return super(ShoppingCart, self).create(vals)

  def write(self, vals):
    if 'state' in vals:
      new_state = vals.get('state')
      if new_state in ['shipping', 'done']:
        for rec in self:
          if any(picking.state != 'done' for picking in rec.picking_ids):
            raise UserError('Bạn không thể chuyển sang trạng thái này khi Phiếu xuất kho chưa Hoàn thành!')
    return super(ShoppingCart, self).write(vals)

  def action_awaiting_confirmation(self):
    for rec in self:
      rec.state = 'awaiting_confirmation'
      rec.message_post(body='Đơn hàng đang chờ được xác nhận')

  def action_draft(self):
    for rec in self:
      rec.state = 'draft'
      rec.message_post(body='Đơn hàng được đặt lại về nháp')

  # gửi mail confirm
  def action_confirm(self):
    template = self.env.ref('om_sales.email_template_cart_confirm', raise_if_not_found=False)
    for rec in self:
      rec.state = 'confirmed'
      rec.message_post(body='Đơn hàng đã được xác nhận')

      # Tạo Phiếu Xuất Kho Tự Động
      internal_loc, customer_loc = rec._get_default_locations()
      if not rec.picking_ids:
        picking = self.env['sm.stock.picking'].create(
          {'picking_type': 'out', 'location_id': internal_loc.id, 'location_dest_id': customer_loc.id,
            'cart_id': rec.id, 'customer_name': rec.customer_name, })
        for line in rec.cart_line_ids:
          if line.product_id:
            self.env['sm.stock.move'].create(
              {'name': f"Xuất bán {line.product_id.name}", 'product_id': line.product_id.id, 'quantity': line.quantity,
                'picking_id': picking.id, })

      if template and rec.customer_email:
        template.with_context(lang=self.env.lang).send_mail(rec.id, force_send=True, raise_exception=True)

  # mail hủy
  def action_cancel(self):
    template = self.env.ref('om_sales.email_template_cart_cancel', raise_if_not_found=False)
    for rec in self:
      rec.state = 'cancel'
      rec.message_post(body='Đơn hàng đã bị hủy')

      # Huỷ các phiếu kho liên quan nếu nó chưa hoàn thành
      for picking in rec.picking_ids:
        if picking.state != 'done':
          picking.action_cancel()

      if template and rec.customer_email:
        template.send_mail(rec.id, force_send=True)

  # def action_shipping(self):
  #   template = self.env.ref('om_sales.email_template_cart_shipping', raise_if_not_found=False)
  #   for rec in self:
  #     # Kiểm tra xem có Phiếu Kho nào chưa hoàn thành không
  #     if any(picking.state != 'done' for picking in rec.picking_ids):
  #       raise UserError('Bạn không thể chuyển sang Đang giao hàng khi Phiếu xuất kho chưa Hoàn thành!')
  #
  #     rec.state = 'shipping'
  #     rec.message_post(body='Đơn hàng đang giao')
  #     if template and rec.customer_email:
  #       template.send_mail(rec.id, force_send=True)

  def action_shipping(self):
    for rec in self:
      if any(picking.state != 'done' for picking in rec.picking_ids):
        raise UserError('Bạn không thể chuyển sang Đang giao hàng khi Phiếu xuất kho chưa Hoàn thành!')

      rec.state = 'shipping'
      rec.message_post(body='Đơn hàng đang giao')

  def action_return(self):
    for rec in self:
      rec.state = 'returned'
      rec.message_post(body='Đơn hàng bị Khách trả lại. Hệ thống đang tạo Phiếu Nhập Kho (Return).')

      # Tạo Phiếu Nhập Kho hoàn trả (Return)
      internal_loc, customer_loc = rec._get_default_locations()
      picking = self.env['sm.stock.picking'].create(
        {'picking_type': 'in', 'location_id': customer_loc.id, 'location_dest_id': internal_loc.id, 'cart_id': rec.id,
          'customer_name': rec.customer_name, })
      for line in rec.cart_line_ids:
        if line.product_id:
          self.env['sm.stock.move'].create(
            {'name': f"Khách trả {line.product_id.name}", 'product_id': line.product_id.id, 'quantity': line.quantity,
              'picking_id': picking.id, })

  # mail hoàn thành
  # def action_done(self):
  #   template = self.env.ref('om_sales.email_template_cart_done', raise_if_not_found=False)
  #   for rec in self:
  #     rec.state = 'done'
  #     rec.message_post(body='Đơn hàng đã giao thành công')
  #     if template and rec.customer_email:
  #       template.send_mail(rec.id, force_send=True)

  def action_done(self):
    for rec in self:
      rec.state = 'done'
      rec.message_post(body='Đơn hàng đã giao thành công')

  def action_print_order(self):
    self.ensure_one()
    return self.env.ref('om_sales.action_report_shopping_cart').report_action(self)

  def _append_to_consolidated_excel(self):
    pass  # No longer needed if we regenerate from scratch, but keep the definition so the controller doesn't crash

  def action_download_consolidated_excel(self):
    from odoo.exceptions import UserError
    orders = self.env['sm.shopping.cart'].search([('payment_type', '=', 'bank_transfer')])

    if not orders:
      raise UserError("Hiện tại chưa có đơn hàng nào thanh toán qua Chuyển khoản / QR Code.")

    attachment = self._get_or_create_consolidated_attachment()
    if attachment:
      return {'type': 'ir.actions.act_url', 'url': f'/web/content/{attachment.id}?download=true', 'target': 'new', }
    raise UserError("Đã xảy ra lỗi không xác định khi tạo file Excel. Không thể lấy dữ liệu.")

  def _get_or_create_consolidated_attachment(self):
    """Lấy attachment Excel tổng hợp. Tạo mới hoàn toàn từ toàn bộ đơn chuyển khoản."""
    import base64
    from io import BytesIO
    MASTER_FILENAME = "DonHang_ChuyenKhoan_TongHop.xlsx"
    company = self.env.company
    Attachment = self.env['ir.attachment'].sudo()

    orders = self.env['sm.shopping.cart'].search([('payment_type', '=', 'bank_transfer')], order='date_order desc')
    if not orders:
      return None

    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DonHang_ChuyenKhoan"

    header_font = Font(bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    headers = ["Mã đơn", "Ngày đặt", "Tên KH", "SĐT", "Email", "Địa chỉ", "Sản phẩm", "Tổng tiền hàng (VNĐ)",
               "Giảm giá (VNĐ)", "Mã giảm giá", "Thành tiền CK (VNĐ)"]

    ws.cell(1, 1, "DANH SÁCH ĐƠN HÀNG THANH TOÁN CHUYỂN KHOẢN").font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for col, h in enumerate(headers, 1):
      c = ws.cell(2, col, h)
      c.font = header_font
      c.border = thin_border
      c.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    next_row = 3
    for order in orders:
      products_text = "; ".join(
        [f"{line.product_id.name} x{line.quantity}" if line.product_id else "" for line in order.cart_line_ids]) or "-"
      coupon_names = ", ".join(order.coupon_ids.mapped('name')) if order.coupon_ids else (
        order.coupon_id.name if order.coupon_id else "-")

      row_data = [order.name, order.date_order.strftime('%d/%m/%Y %H:%M') if order.date_order else "",
                  order.customer_name or "", order.customer_phone or "", order.customer_email or "",
                  (order.customer_address or "").replace("\n", " "), products_text, f"{order.subtotal_price:,.0f}",
                  f"-{order.discount_amount:,.0f}", coupon_names, f"{order.total_price:,.0f}", ]

      for col, val in enumerate(row_data, 1):
        c = ws.cell(next_row, col, val)
        c.border = thin_border
      next_row += 1

    # Cột rộng
    for i, w in enumerate([15, 16, 22, 14, 28, 35, 40, 18, 16, 25, 20], 1):
      ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_data = base64.b64encode(buffer.getvalue())

    master = Attachment.search(
      [('res_model', '=', 'res.company'), ('res_id', '=', company.id), ('name', '=', MASTER_FILENAME), ], limit=1)

    vals = {'name': MASTER_FILENAME, 'datas': file_data, 'res_model': 'res.company', 'res_id': company.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', }

    if master:
      master.write(vals)
    else:
      master = Attachment.create(vals)

    return master

  # xử lí dữ liệu để đẩy lên dashboard
  @api.model
  def get_owl_dashboard_data(self, date_filter="all", date_start=None, date_end=None):
    from dateutil.relativedelta import relativedelta
    today = fields.Date.context_today(self)

    domain = []
    if date_filter == 'today':
      domain.append(('date_order', '>=', today))
    elif date_filter == 'this_month':
      start_of_month = today.replace(day=1)
      domain.append(('date_order', '>=', start_of_month))
    elif date_filter == 'this_year':
      start_of_year = today.replace(month=1, day=1)
      domain.append(('date_order', '>=', start_of_year))
    elif date_filter == 'custom':
      if date_start:
        domain.append(('date_order', '>=', f"{date_start} 00:00:00"))
      if date_end:
        domain.append(('date_order', '<=', f"{date_end} 23:59:59"))

    # KPI: Total Products (Laptops available)
    total_laptops = self.env['sm.sanpham'].search_count([])

    # KPI: Order Counts by State
    all_orders = self.env['sm.shopping.cart'].search(domain)
    order_count = len(all_orders)
    # hiển thị các tình trạng đơn
    pending_count = len(all_orders.filtered(lambda o: o.state in ['draft', 'awaiting_confirmation', 'shipping']))
    done_count = len(all_orders.filtered(lambda o: o.state == 'done'))
    cancel_count = len(all_orders.filtered(lambda o: o.state == 'cancel'))
    # tính tổng giá trị các đơn hàng
    # KPI: Total Revenue (from done or confirmed/shipping non-cancelled orders)
    revenue_orders = all_orders.filtered(
      lambda o: o.state not in ['draft', 'cancel', 'awaiting_confirmation', 'returned'])
    total_revenue = sum(revenue_orders.mapped('total_price'))

    # MoM Revenue Comparison
    start_of_this_month = today.replace(day=1)
    start_of_last_month = (start_of_this_month - relativedelta(months=1))
    # This month revenue
    this_month_orders = self.env['sm.shopping.cart'].search(
      [('state', 'not in', ['draft', 'cancel', 'awaiting_confirmation', 'returned']),
       ('date_order', '>=', start_of_this_month)])
    this_month_revenue = sum(this_month_orders.mapped('total_price'))

    # Last month revenue
    last_month_orders = self.env['sm.shopping.cart'].search(
      [('state', 'not in', ['draft', 'cancel', 'awaiting_confirmation', 'returned']),
       ('date_order', '>=', start_of_last_month), ('date_order', '<', start_of_this_month)])
    last_month_revenue = sum(last_month_orders.mapped('total_price'))

    mom_growth = 0
    if last_month_revenue > 0:
      mom_growth = ((this_month_revenue - last_month_revenue) / last_month_revenue) * 100
    elif this_month_revenue > 0:
      mom_growth = 100  # Infinity, but represented as 100%

    # Chart Data
    # Revenue by Month (Current Year)
    start_of_year = today.replace(month=1, day=1)
    year_orders = self.env['sm.shopping.cart'].search(
      [('state', 'not in', ['draft', 'cancel', 'awaiting_confirmation', 'returned']),
       ('date_order', '>=', start_of_year)])

    revenue_by_month = {}
    for m in range(1, 13):
      revenue_by_month[str(m)] = 0

    for o in year_orders:
      if o.date_order:
        month_key = str(o.date_order.month)
        revenue_by_month[month_key] += o.total_price

    # Orders by Day (Current Month)
    month_orders = self.env['sm.shopping.cart'].search([('date_order', '>=', start_of_this_month)])
    days_in_month = (today.replace(day=28) + relativedelta(days=4)).replace(day=1) - relativedelta(days=1)
    orders_by_day = {}
    for d in range(1, days_in_month.day + 1):
      orders_by_day[str(d)] = 0

    for o in month_orders:
      if o.date_order:
        day_key = str(o.date_order.day)
        orders_by_day[day_key] += 1

    # Top Selling Products (Lifetime or filtered? Let's use filtered for relevance)
    order_lines = self.env['sm.shopping.cart.line'].search([('cart_id', 'in', all_orders.ids)])
    product_sales = {}
    for line in order_lines:
      if line.product_id:
        if line.product_id.id not in product_sales:
          product_sales[line.product_id.id] = {'name': line.product_id.name, 'code': line.product_id.code, 'qty': 0,
                                               'revenue': 0}
        product_sales[line.product_id.id]['qty'] += line.quantity
        product_sales[line.product_id.id]['revenue'] += line.price_subtotal

    top_products = sorted(list(product_sales.values()), key=lambda x: x['qty'], reverse=True)[:5]

    # Top Discount Codes (Based on all applied coupons in the filtered orders)
    coupon_usage = {}
    for o in all_orders:
      # We check both coupon_id and coupon_ids
      coupons_to_check = o.coupon_ids if o.coupon_ids else (o.coupon_id if o.coupon_id else None)
      if not coupons_to_check:
        continue

      # If coupons_to_check is not a list (it's a recordset)
      for c in coupons_to_check:
        if c.id not in coupon_usage:
          coupon_usage[c.id] = {'name': c.name, 'count': 0, 'discount_value': c.discount_value,
                                'discount_type': c.discount_type}
        coupon_usage[c.id]['count'] += 1

    top_coupons = sorted(list(coupon_usage.values()), key=lambda x: x['count'], reverse=True)[:5]

    return {'total_laptops': total_laptops, 'order_count': order_count, 'pending_count': pending_count,
            'done_count': done_count, 'cancel_count': cancel_count, 'total_revenue': total_revenue,
            'this_month_revenue': this_month_revenue, 'last_month_revenue': last_month_revenue,
            'mom_growth': round(mom_growth, 2), 'currency_symbol': self.env.company.currency_id.symbol or 'VNĐ',
            # tạo thành 12 điểm theo trục x
            'chart_revenue_labels': [f"T{m}" for m in range(1, 13)], #       lấy doanh thu tương ứng trục y
            'chart_revenue_data': [revenue_by_month[str(m)] for m in range(1, 13)],
            'chart_orders_labels': [f"Ngày {d}" for d in range(1, days_in_month.day + 1)],
            'chart_orders_data': [orders_by_day[str(d)] for d in range(1, days_in_month.day + 1)],
            'top_products': top_products, 'top_coupons': top_coupons}


class ShoppingCartLine(models.Model):
  _name = 'sm.shopping.cart.line'
  _description = 'Chi tiết đơn hàng'

  cart_id = fields.Many2one('sm.shopping.cart', string='Đơn hàng', ondelete='cascade')
  product_id = fields.Many2one('sm.sanpham', string='Tên sản phẩm')
  quantity = fields.Integer(string='Số lượng', default=1)
  price_unit = fields.Float(string='Đơn giá', required=True, default=0.0)
  price_subtotal = fields.Float(string='Thành tiền', compute='_compute_subtotal', store=True)

  @api.onchange('product_id')
  def _onchange_product_id(self):
    if self.product_id:
      self.price_unit = self.product_id.current_discounted_price

  lot_names = fields.Char(string='Số Serial', compute='_compute_lot_names')

  def _compute_lot_names(self):
    for line in self:
      # Tìm các phiếu giao hàng đã hoàn thành của đơn hàng này
      done_pickings = line.cart_id.picking_ids.filtered(lambda p: p.state == 'done' and p.picking_type == 'out')

      # Tìm các dòng di chuyển kho thuộc về sản phẩm này trong các phiếu trên
      moves = self.env['sm.stock.move'].search(
        [('picking_id', 'in', done_pickings.ids), ('product_id', '=', line.product_id.id), ('state', '=', 'done')])

      # Lấy danh sách tên Số Serial đã chọn/tạo
      lot_names = []
      for move in moves:
        for m_line in move.move_line_ids:
          if m_line.lot_id:
            lot_names.append(m_line.lot_id.name)

      line.lot_names = ", ".join(lot_names) if lot_names else ""

  @api.depends('quantity', 'price_unit')
  def _compute_subtotal(self):
    for line in self:
      line.price_subtotal = line.quantity * line.price_unit